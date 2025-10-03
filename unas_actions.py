import os
import subprocess
import json
from datetime import datetime, timedelta, date
from typing import Dict, Optional
import requests
import xml.etree.ElementTree as ET
from copy import deepcopy
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# -----------------------------
# Setup
# -----------------------------
load_dotenv()
UNAS_API_BASE = os.getenv('UNAS_API_BASE')
UNAS_API_KEY = os.getenv("UNAS_API_KEY")
SESSION_TIMEOUT = 20  # sec


def data_dir_with_filename(fname: str) -> str:
    os.makedirs("data", exist_ok=True)
    return f"data/{fname}"


# -----------------------------
# Business rules
# -----------------------------
ALLOWED_CUSTOMER_GROUPS = {"", "Alapértelmezett", "SAP9-Törzsvásárló"}

SKIP_ITEM_NAME_SUBSTRINGS = (
    "szállítási költség",
    "utánvét kezelési költség",
)

ORDER_COLUMNS: Dict[str, str] = {
    "Order_Id": "Id",
    "Order_Key": "Key",
    "Order_Date": "Date",
    "Order_DateMod": "DateMod",
    "Order_Status": "Status",
    "Order_StatusID": "StatusID",
    "Order_Currency": "Currency",
    "Order_SumPriceGross": "SumPriceGross",
    "Order_Referer": "Referer",
    "Order_CustomerEmail": "Customer/Email",
    "Order_CustomerName": "Customer/Contact/Name",
    "Order_CustomerLang": "Customer/Contact/Lang",
    "Order_CustomerGroup": "Customer/Group/Name",
    "Order_InvoiceZIP": "Customer/Addresses/Invoice/ZIP",
    "Order_InvoiceCity": "Customer/Addresses/Invoice/City",
    "Order_InvoiceCountry": "Customer/Addresses/Invoice/Country",
    "Order_ShippingZIP": "Customer/Addresses/Shipping/ZIP",
    "Order_ShippingCity": "Customer/Addresses/Shipping/City",
    "Order_ShippingCountry": "Customer/Addresses/Shipping/Country",
    "Order_PaymentName": "Payment/Name",
    "Order_PaymentType": "Payment/Type",
    "Order_PaymentStatus": "Payment/Status",
    "Order_Paid": "Payment/Paid",
    "Order_Unpaid": "Payment/Unpaid",
    "Order_UTM_Source": "UTM/Source",
    "Order_UTM_Medium": "UTM/Medium",
    "Order_UTM_Campaign": "UTM/Campaign",
    "Order_UTM_Content": "UTM/Content",
}


# -----------------------------
# XML helpers
# -----------------------------
def _xml(params_dict: dict) -> str:
    root = ET.Element("Params")
    for k, v in params_dict.items():
        e = ET.SubElement(root, k)
        e.text = str(v)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True).decode("utf-8")


def txt(el: ET.Element, path: str) -> str:
    f = el.find(path)
    return f.text.strip() if f is not None and f.text is not None else ""


def write_response_xml_file(string: str, fname: str) -> None:
    with open(data_dir_with_filename(fname), "w", encoding="utf-8") as f:
        f.write(string)


# -----------------------------
# UNAS auth + calls
# -----------------------------
def unas_login(api_key: str) -> str:
    url = f"{UNAS_API_BASE}/login"
    body = _xml({"ApiKey": api_key, "WebshopInfo": "true"})
    headers = {"Content-Type": "application/xml"}
    resp = requests.post(url, data=body.encode("utf-8"), headers=headers, timeout=SESSION_TIMEOUT)
    resp.raise_for_status()
    tree = ET.fromstring(resp.text)
    token_el = tree.find("Token")
    if token_el is None or not token_el.text:
        raise RuntimeError(f"Token not found. Response: {resp.text[:500]}")
    return token_el.text.strip()


def set_token(token: str) -> None:
    with open("token.txt", "w", encoding="utf-8") as f:
        f.write(token)


def get_token() -> str:
    with open("token.txt", "r", encoding="utf-8") as f:
        return f.read().strip()


def get_unas_token() -> None:
    token = unas_login(UNAS_API_KEY)
    print(f"Token OK: {token[:8]}...")  # shortened
    set_token(token)


def unas_call(method: str, params: dict) -> ET.Element:
    url = f"{UNAS_API_BASE}/{method}"
    body = _xml(params)
    headers = {
        "Content-Type": "application/xml",
        "Authorization": f"Bearer {get_token()}",
    }
    resp = requests.post(url, data=body.encode("utf-8"), headers=headers, timeout=SESSION_TIMEOUT)
    resp.raise_for_status()
    return ET.fromstring(resp.text)


def get_all_orders(date_start: str, date_end: str, batch_size: int = 500, max_pages: int = 2000) -> str:
    """
    Összes rendelés lekérése a megadott intervallumban, 500-as lapozással.
    Visszatérés: egyetlen <Orders> XML szövegben (összefűzve).
    """
    combined_chunks = []
    start = 0
    pages = 0

    while True:
        params = {
            "DateStart": date_start,
            "DateEnd": date_end,
            "LimitNum": batch_size,
            "LimitStart": start,
        }
        orders_elem = unas_call("getOrder", params)
        count = len(orders_elem.findall(".//Order"))

        chunk_xml = ET.tostring(orders_elem, encoding="utf-8", xml_declaration=True).decode("utf-8")
        if count > 0:
            combined_chunks.append(chunk_xml)

        print(f"Fetched page start={start} size={batch_size}, orders={count}")
        pages += 1
        if count < batch_size or pages >= max_pages:
            break
        start += batch_size

    if not combined_chunks:
        return '<?xml version="1.0" encoding="utf-8"?><Orders></Orders>'
    return combine_orders_xml_strings(*combined_chunks)


# -----------------------------
# Flatten helpers
# -----------------------------
def _flatten_element(elem: ET.Element, base_path: str = "") -> dict:
    out = {}
    if elem.attrib:
        for k, v in elem.attrib.items():
            out[(base_path + "/@" + k).strip("/")] = v
    children = list(elem)
    if not children:
        out[base_path.strip("/")] = (elem.text or "").strip()
        return out
    by_tag = {}
    for child in children:
        by_tag.setdefault(child.tag, []).append(child)
    for tag, nodes in by_tag.items():
        if len(nodes) == 1:
            child_path = f"{base_path}/{tag}".strip("/")
            out.update(_flatten_element(nodes[0], child_path))
        else:
            for idx, node in enumerate(nodes, start=1):
                child_path = f"{base_path}/{tag}[{idx}]".strip("/")
                out.update(_flatten_element(node, child_path))
    return out


def _should_skip_item_by_name(item_elem: ET.Element) -> bool:
    name = txt(item_elem, "Name").lower()
    return any(substr in name for substr in SKIP_ITEM_NAME_SUBSTRINGS)


# -----------------------------
# Optional: Combine XMLs -> single <Orders>
# -----------------------------
def combine_orders_xml_strings(*xml_strings: str) -> str:
    combined = ET.Element("Orders")
    for x in xml_strings:
        if not x or not x.strip():
            continue
        root = ET.fromstring(x)
        for order in root.findall(".//Order"):
            combined.append(deepcopy(order))
    return ET.tostring(combined, encoding="utf-8", xml_declaration=True).decode("utf-8")


# -----------------------------
# XML -> DataFrame utilities (ORDER-LEVEL ONLY)
# -----------------------------
def xml_string_to_dataframe(xml_text: str) -> pd.DataFrame:
    """
    Rendelésenként EGY soros táblázatot ad vissza.
    Csak az ORDER_COLUMNS mezőit tölti (Customer Group szűréssel),
    NINCSENEK Item sorok, LineNo, Item_* oszlopok.
    """
    root = ET.fromstring(xml_text)
    rows = []
    for o in root.findall(".//Order"):
        group_name = txt(o, "Customer/Group/Name")
        if group_name not in ALLOWED_CUSTOMER_GROUPS:
            continue
        order_ctx = {col_name: txt(o, xpath) for col_name, xpath in ORDER_COLUMNS.items()}
        rows.append(order_ctx)
    if not rows:
        return pd.DataFrame(columns=list(ORDER_COLUMNS.keys()))
    return pd.DataFrame(rows, columns=list(ORDER_COLUMNS.keys()))


def xml_file_to_dataframe(xml_path: str) -> pd.DataFrame:
    xml_text = open(xml_path, "r", encoding="utf-8").read()
    return xml_string_to_dataframe(xml_text)


# -----------------------------
# DataFrame -> new Excel utilities
# -----------------------------
def write_dataframe_to_new_excel(df: pd.DataFrame, out_xlsx: str, sheet_name: str = "OrderItems_ALL") -> str:
    os.makedirs(os.path.dirname(out_xlsx) or ".", exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xlw:
        df.to_excel(xlw, sheet_name=sheet_name, index=False)
    return out_xlsx


def export_xml_file_to_excel_one_sheet(xml_path: str, out_xlsx: Optional[str] = None,
                                       sheet_name: str = "OrderItems_ALL") -> str:
    if out_xlsx is None:
        out_xlsx = f"{os.path.splitext(xml_path)[0]}.xlsx"
    df = xml_file_to_dataframe(xml_path)
    return write_dataframe_to_new_excel(df, out_xlsx, sheet_name=sheet_name)


# --- Compatibility shim (so old code continues to work) ---
def xml_to_excel_one_sheet(xml_path: str, out_xlsx: Optional[str] = None) -> str:
    return export_xml_file_to_excel_one_sheet(xml_path, out_xlsx, sheet_name="OrderItems_ALL")


# -----------------------------
# Helpers to manage batches in Excel (top-insert flow)
# -----------------------------
def _open_or_init_wb_with_header(xlsx_path: str, sheet_name: str, header_cols: list):
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 1:
                ws.append(header_cols)
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(header_cols)
    else:
        os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(header_cols)
    return wb, ws


def _find_batch_bounds(ws, label: str) -> Optional[tuple]:
    target = f"Day: {label}"
    start = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val == target:
            start = r
            break
    if start is None:
        return None
    next_batch = None
    for r in range(start + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.startswith("Day: "):
            next_batch = r
            break
    end = (next_batch - 1) if next_batch else ws.max_row
    return start, end


def delete_batch_by_label(xlsx_path: str, sheet_name: str, label: str, header_cols: list) -> bool:
    wb, ws = _open_or_init_wb_with_header(xlsx_path, sheet_name, header_cols)
    bounds = _find_batch_bounds(ws, label)
    if not bounds:
        wb.save(xlsx_path)
        return False
    start, end = bounds
    amount = end - start + 1
    ws.delete_rows(start, amount)
    wb.save(xlsx_path)
    return True


def prepend_batch_to_excel(df: pd.DataFrame, xlsx_path: str, batch_label: str, sheet_name: str = "OrderItems_ALL",
                           spacer_rows: int = 3) -> None:
    header_cols = list(df.columns)
    wb, ws = _open_or_init_wb_with_header(xlsx_path, sheet_name, header_cols)
    n_rows = len(df)
    total_to_insert = 1 + n_rows + 1 + spacer_rows  # label + data + summary + spacer
    ws.insert_rows(idx=2, amount=total_to_insert)
    ws.cell(row=2, column=1, value=f"Day: {batch_label}")
    start_data_row = 3
    for i, row_vals in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=start_data_row + i, column=col_idx, value=val)
    if "Order_Key" in df.columns:
        orders_count = int(df["Order_Key"].nunique())
    elif "Order_Id" in df.columns:
        orders_count = int(df["Order_Id"].nunique())
    else:
        orders_count = int(len(df))
    summary_row_idx = start_data_row + n_rows
    ws.cell(row=summary_row_idx, column=1, value="Orders on day:")
    ws.cell(row=summary_row_idx, column=2, value=orders_count)
    wb.save(xlsx_path)


# -----------------------------
# Weekly ranges helpers
# -----------------------------
def weekly_ranges_back(months=1, fmt="%Y.%m.%d") -> list:
    today = date.today()
    this_week_monday = today - timedelta(days=today.weekday())
    prev_monday = this_week_monday - timedelta(days=7)
    prev_sunday = this_week_monday - timedelta(days=1)
    max_weeks = months * 4 + 2
    ranges = []
    weeks_ago = 1
    while weeks_ago <= max_weeks:
        ranges.append({
            "weeks_ago": weeks_ago,
            "start": prev_monday.strftime(fmt),
            "end": prev_sunday.strftime(fmt),
        })
        prev_monday -= timedelta(days=7)
        prev_sunday -= timedelta(days=7)
        weeks_ago += 1
    return ranges


def save_week_ranges() -> None:
    os.makedirs("data", exist_ok=True)
    json.dump(weekly_ranges_back(), open("weekly_ranges.json", "w", encoding="utf-8"))


def get_week_ranges() -> dict:
    data = json.load(open("weekly_ranges.json", encoding="utf-8"))
    weeks = {}
    for line in data:
        weeks[line["weeks_ago"]] = f"{line['start']}-{line['end']}"
    return weeks


# -----------------------------
# NEW: Month-based workbook builders
# -----------------------------
def month_sheet_name(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


def ensure_wb_and_sheet(xlsx_path: str, sheet_name: str, header_cols: list):
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 1:
                ws.append(header_cols)
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(header_cols)
    else:
        os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(header_cols)
    return wb, ws


def append_week_block(ws, df: pd.DataFrame, label: str, spacer_rows: int = 3):
    """
    Hozzáfűz egy heti blokkot a lap ALJÁRA:
    - 1 sor: "Batch: YYYY.MM.DD-YYYY.MM.DD"
    - df tartalma (fejléc NÉLKÜL)
    - 1 sor: "Orders in week:" <egyedi Order_Id>
    - spacer_rows db ÜRES sor (valóban üresen)
    """
    insert_at = (ws.max_row or 1) + 1

    # Címke
    ws.cell(row=insert_at, column=1, value=f"Week: {label}")

    # Adatsorok
    for i, row_vals in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=insert_at + 1 + i, column=c_idx, value=val)

    # Összegzés
    orders_count = int(df["Order_Id"].nunique() if "Order_Id" in df.columns else len(df))
    summary_row = insert_at + 1 + len(df)
    ws.cell(row=summary_row, column=1, value="Orders in week:")
    ws.cell(row=summary_row, column=2, value=orders_count)

    # ÜRES sorok beszúrása biztosan
    for idx in range(max(0, spacer_rows)):
        ws.cell(row=summary_row + idx + 1, column=2, value="")


def week_months_covered(start_dt: date, end_dt: date) -> list[date]:
    """
    Visszaadja a hét által érintett hónapok első napjait (egy vagy kettő).
    Ha a hét átlóg, mindkét hónap szerepeljen.
    """
    months = set()
    cur = start_dt
    while cur <= end_dt:
        months.add(date(cur.year, cur.month, 1))
        cur += timedelta(days=1)
    return sorted(months)


def parse_ymd_dot(s: str) -> date:
    return datetime.strptime(s, "%Y.%m.%d").date()


def weekly_ranges_between(start_dt: date, end_dt: date) -> list[tuple[date, date]]:
    """
    Hétfő–vasárnap bontás a megadott intervallumra (a szélek vágva a megadott ablakhoz).
    """
    first_monday = start_dt - timedelta(days=start_dt.weekday())
    ranges = []
    cur_start = first_monday
    while cur_start <= end_dt:
        cur_end = cur_start + timedelta(days=6)
        s = max(cur_start, start_dt)
        e = min(cur_end, end_dt)
        ranges.append((s, e))
        cur_start += timedelta(days=7)
    return ranges


def build_monthly_workbook_for_previous_weeks(
        months_back: int = 3,
        out_xlsx: str = "data/orders_by_month.xlsx",
        spacer_rows: int = 3
) -> str:
    """
    Az elmúlt 'months_back' hónapra:
      - hónaponként külön sheet (YYYY-MM),
      - a hónapban szereplő hetek blokkja egymás alatt, 3 üres sorral elválasztva,
      - ha a hét két hónapot érint, a blokk mindkét hónap lapjára bemásolva.
    A táblázat rendelés-szintű (xml_string_to_dataframe) — 1 sor / rendelés.
    """
    today = date.today()
    cur_month_first = date(today.year, today.month, 1)

    start_month_year = cur_month_first.year
    start_month = cur_month_first.month - months_back
    while start_month <= 0:
        start_month += 12
        start_month_year -= 1
    window_start = date(start_month_year, start_month, 1)
    window_end = today  # ha lezárt hónapig kellene, állítsd a hónap utolsó napjára

    weeks = weekly_ranges_between(window_start, window_end)
    header_cols = list(ORDER_COLUMNS.keys())

    for (ws_start, ws_end) in weeks:
        start_s = ws_start.strftime("%Y.%m.%d")
        end_s = ws_end.strftime("%Y.%m.%d")

        xml_week = get_all_orders(date_start=start_s, date_end=end_s)
        df_week = xml_string_to_dataframe(xml_week)  # 1 sor / rendelés

        if df_week.empty:
            continue

        label = f"{start_s}-{end_s}"
        months_hit = week_months_covered(ws_start, ws_end)

        for month_first_day in months_hit:
            sheet = month_sheet_name(month_first_day)
            wb, ws = ensure_wb_and_sheet(out_xlsx, sheet, header_cols)
            append_week_block(ws, df_week, label=label, spacer_rows=spacer_rows)
            wb.save(out_xlsx)

    print(f"Monthly workbook ready: {out_xlsx}")
    return out_xlsx


# -----------------------------
# Fetchers (file-per-run exports)
# -----------------------------
def fetch_today_orders_and_export_excel(day_else: Optional[str] = None) -> str:
    day = day_else or datetime.now().strftime("%Y.%m.%d")
    response = get_all_orders(date_start=day, date_end=day)
    fname_xml = f"today.xml"
    write_response_xml_file(response, fname_xml)
    src_xml = f"data/{fname_xml}"
    out_xlsx = f"data/today_{day}.xlsx"
    export_xml_file_to_excel_one_sheet(src_xml, out_xlsx)
    print(f"Export kész: {out_xlsx}")
    return out_xlsx


def fetch_previous_months_orders_and_export_excel() -> None:
    """(Opcionális régi funkció) Heti fájlok külön xlsx-be."""
    save_week_ranges()
    for week in get_week_ranges().values():
        start_date, end_date = week.split('-')
        fname_xml = f"week_{start_date}-{end_date}.xml"
        write_response_xml_file(get_all_orders(start_date, end_date), fname_xml)
        src_xml = f"data/{fname_xml}"
        out_xlsx = f"data/week_{start_date}-{end_date}.xlsx"
        export_xml_file_to_excel_one_sheet(src_xml, out_xlsx)
        print("Export ready:", out_xlsx, src_xml)


# -----------------------------
# Daily job (18:00): optional example
# -----------------------------
def daily_summary_orders_to_excel(output_path: str = "data/orders_main.xlsx", sheet_name: str = "OrderItems_ALL",
                                  spacer_rows: int = 3) -> None:
    today_str = datetime.now().strftime("%Y.%m.%d")
    yday_str = (date.today() - timedelta(days=1)).strftime("%Y.%m.%d")
    xml_today = get_all_orders(date_start=today_str, date_end=today_str)
    xml_yday = get_all_orders(date_start=yday_str, date_end=yday_str)
    write_response_xml_file(xml_today, "today.xml")
    write_response_xml_file(xml_yday, "yesterday.xml")
    df_today = xml_string_to_dataframe(xml_today)
    df_yday = xml_string_to_dataframe(xml_yday)
    header_cols = list(df_today.columns) if len(df_today.columns) >= len(df_yday.columns) else list(df_yday.columns)
    _open_or_init_wb_with_header(output_path, sheet_name, header_cols)[0].save(output_path)
    deleted = delete_batch_by_label(output_path, sheet_name, yday_str, header_cols)
    if deleted:
        print(f"• Removed previous partial day for {yday_str}")
    prepend_batch_to_excel(df_yday, output_path, batch_label=yday_str, sheet_name=sheet_name, spacer_rows=spacer_rows)
    prepend_batch_to_excel(df_today, output_path, batch_label=today_str, sheet_name=sheet_name, spacer_rows=spacer_rows)
    print(f"✔ Rotated batches. Top = TODAY({today_str}), below = YESTERDAY(full {yday_str}). File: {output_path}")

def upload_excel_files_into_google_cloud() -> None:
    subprocess.run(["python", "google_cloud_actions.py"], capture_output=True, text=True)

    print("Google Cloud Actions Upload Complete")
# -----------------------------
# Main
# -----------------------------
def main() -> None:
    # --- get UNAS token
    get_unas_token()

    # --- napi külön export:
    fetch_today_orders_and_export_excel()

    # --- Havi könyv építése az elmúlt 3 hónapra:
    # build_monthly_workbook_for_previous_weeks(
    #     months_back=3,
    #     out_xlsx="data/orders_by_month.xlsx",
    #     spacer_rows=10
    # )

    # --- call google_cloud_actions.py
    # upload_excel_files_into_google_cloud()

    # --- napi összefoglaló top-insert megoldással:
    daily_summary_orders_to_excel()

if __name__ == "__main__":
    main()
