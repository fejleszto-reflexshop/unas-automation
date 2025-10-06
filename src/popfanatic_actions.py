import os
import json
import time
import requests
from dotenv import load_dotenv
from datetime import datetime, timedelta, date
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

load_dotenv()

SHOP_NAME     = (os.getenv("POPFANATIC_SHOP_NAME") or "").strip()
CLIENT_ID     = (os.getenv("POPFANATIC_CLIENT_ID") or "").strip()
CLIENT_SECRET = (os.getenv("POPFANATIC_CLIENT_SECRET") or "").strip()

TOKEN_URL = (os.getenv("POPFANATIC_TOKEN_URL") or f"").strip()
API_BASE  = (os.getenv("POPFANATIC_API_URL")  or "").strip()

DATA_DIR = "../data"
DEFAULT_MAIN_XLSX = os.path.join(DATA_DIR, "orders_popfanatic_main.xlsx")
DEFAULT_SHEET = "Orders_ALL"

# -----------------------------
# FS helpers
# -----------------------------
def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)

# -----------------------------
# Auth + API calls
# -----------------------------
def get_access_token() -> tuple:
    payload = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
    }
    headers = {"Content-Type": "application/json", "Accept": "application/json"}
    r = requests.post(TOKEN_URL, headers=headers, json=payload, timeout=30)

    if r.status_code != 200:
        raise RuntimeError(f"Token error {r.status_code}: {r.text}")

    data = r.json()
    return data["access_token"], data.get("token_type", "Bearer")

def get_orders(access_token, token_type, page=0, limit=200, extra_params=None) -> dict:
    """
    Fetch one page of orders. Flexible envelope handling.
    """
    params = {"page": page, "limit": limit, "full": 0}
    if extra_params:
        params.update(extra_params)

    url = f"{API_BASE}/orders"
    headers = {
        "Authorization": f"{token_type} {access_token}",
        "Accept": "application/json",
    }

    for attempt in range(4):
        r = requests.get(url, headers=headers, params=params, timeout=30)
        if r.status_code == 200:
            return r.json()
        if r.status_code in (429, 500, 502, 503, 504):
            time.sleep(attempt or 1)
            continue

        ct = r.headers.get("Content-Type", "")
        body = r.text if "application/json" not in ct else r.json()
        raise RuntimeError(f"Orders error {r.status_code}: {body}")

    raise RuntimeError("Orders error: retry limit exceeded")

def get_order_by_id(access_token, token_type, order_id) -> dict:
    url = f"{API_BASE}/orders/{order_id}"
    headers = {
        "Authorization": f"{token_type} {access_token}",
        "Accept": "application/json"
    }

    for attempt in range(4):
        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code == 200:
            return r.json()
        if r.status_code in (429, 500, 502, 503, 504):
            time.sleep(attempt or 1)
            continue
        raise RuntimeError(f"Order error {r.status_code}: {r.text}")

    raise RuntimeError("Order error: retry limit exceeded")

def extract_order_id(item: dict) -> str:
    href = item.get("href") or item.get("_links", {}).get("self", {}).get("href")
    if href:
        h = str(href).rstrip("/")
        h = h.split("?", 1)[0].split("#", 1)[0]
        parts = h.split("/")

        if parts:
            return parts[-1]

    return str(item.get("id"))

# -----------------------------
# Range fetchers
# -----------------------------
def fetch_orders_between(access_token, token_type, start_iso: str, end_iso: str) -> list[dict]:
    """
    Pull **all** orders between [start_iso, end_iso] (inclusive), with pagination.
    Returns list of *stubs* (not the detailed objects).
    """
    page = 0
    limit = 200
    collected = []

    extra = {"createdAtMin": start_iso, "createdAtMax": end_iso}

    while True:
        page_data = get_orders(access_token, token_type, page=page, limit=limit, extra_params=extra)
        items = page_data.get("items") or (page_data.get("response", {}) or {}).get("items", []) or []
        if not items:
            break

        collected.extend(items)
        if len(items) < limit:
            break
        page += 1

    return collected

def fetch_order_details_for_items(access_token, token_type, items: list[dict]) -> list[dict]:
    out = []
    for stub in items:
        oid = extract_order_id(stub)
        detail = get_order_by_id(access_token, token_type, oid)
        out.append(detail)

    return out

# -----------------------------
# Flatten JSON → row dict
# -----------------------------
def _flatten_json(obj, parent_key=""):
    """
    Flatten nested dict/list JSON into a single dict with dotted keys.
    Lists are indexed: 'lines[1].sku', etc.
    """
    rows = {}

    if isinstance(obj, dict):
        for k, v in obj.items():
            nk = f"{parent_key}.{k}" if parent_key else k
            rows.update(_flatten_json(v, nk))
    elif isinstance(obj, list):
        for i, v in enumerate(obj, start=1):
            nk = f"{parent_key}[{i}]"
            rows.update(_flatten_json(v, nk))
    else:
        rows[parent_key] = obj

    return rows

def details_to_dataframe(details: list[dict]) -> pd.DataFrame:
    flat_rows = [_flatten_json(d) for d in details]
    if not flat_rows:
        return pd.DataFrame()

    return pd.DataFrame(flat_rows)

# -----------------------------
# Unique order counter (use innerId if available)
# -----------------------------
PRIORITY_ID_FIELDS = [
    "innerId",  # canonical unique order id
    "id", "order.id", "orderId", "order_id", "number", "incrementId", "code", "Order_Id", "Order_Key"
]
EXCLUDE_ID_KEYWORDS = ("customer", "buyer", "user", "address", "billing", "shipping",
                       "product", "item", "line", "sku", "variant")

def estimate_unique_order_count(df: pd.DataFrame) -> int:
    if df is None or df.empty:
        return 0

    cols = list(df.columns)
    lowered = {c.lower(): c for c in cols}

    # 1) Prefer innerId
    if "innerid" in lowered:
        c = lowered["innerid"]
        return int(pd.Series(df[c]).astype(str).nunique(dropna=True))

    # 2) Other known names
    for pname in PRIORITY_ID_FIELDS:
        if pname.lower() in lowered:
            c = lowered[pname.lower()]
            try:
                return int(pd.Series(df[c]).astype(str).nunique(dropna=True))
            except Exception:
                pass

    # 3) Heuristic id columns
    candidates = []
    for c in cols:
        lc = c.lower()
        if any(k in lc for k in EXCLUDE_ID_KEYWORDS):
            continue
        if lc.endswith("id") or lc.endswith(".id") or lc.endswith("_id"):
            try:
                nun = pd.Series(df[c]).astype(str).nunique(dropna=True)
                candidates.append((nun, c))
            except Exception:
                continue

    if candidates:
        candidates.sort(reverse=True)
        return int(candidates[0][0])

    # 4) Fallback
    return int(len(df))

# -----------------------------
# Excel helpers
# -----------------------------
def _open_or_init_wb_with_header(xlsx_path: str, sheet_name: str, header_cols: list[str]):
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 1:
                ws.append(header_cols)
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(header_cols)
    else:
        os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(header_cols)

    return wb, ws

def _get_existing_header(ws) -> list[str]:
    # Read header from first row
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

def _existing_week_labels_in_sheet(ws) -> set[str]:
    """
    Scan column A for lines like 'Week: YYYY.MM.DD-YYYY.MM.DD' and return the labels.
    """
    labels = set()
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.startswith("Week: "):
            label = val.replace("Week: ", "").strip()
            if label:
                labels.add(label)

    return labels

def _find_batch_bounds(ws, label: str):
    target = f"Day: {label}"
    start = None

    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val == target:
            start = r
            break

    if start is None:
        return None

    next_batch = None
    for r in range(start + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.startswith("Day: "):
            next_batch = r
            break
    end = (next_batch - 1) if next_batch else ws.max_row

    return start, end

def delete_batch_by_label(xlsx_path: str, sheet_name: str, label: str, header_cols: list[str]) -> bool:
    wb, ws = _open_or_init_wb_with_header(xlsx_path, sheet_name, header_cols)
    bounds = _find_batch_bounds(ws, label)

    if not bounds:
        wb.save(xlsx_path)
        return False

    start, end = bounds
    amount = end - start + 1
    ws.delete_rows(start, amount)
    wb.save(xlsx_path)

    return True

def prepend_batch_to_excel(df: pd.DataFrame, xlsx_path: str, batch_label: str,
                           sheet_name: str = DEFAULT_SHEET, spacer_rows: int = 3) -> None:
    header_cols = list(df.columns)
    wb, ws = _open_or_init_wb_with_header(xlsx_path, sheet_name, header_cols)

    n_rows = len(df)
    total_to_insert = 1 + n_rows + 1 + spacer_rows  # label + data + summary + spacer
    ws.insert_rows(idx=2, amount=total_to_insert)
    ws.cell(row=2, column=1, value=f"Day: {batch_label}")
    start_data_row = 3

    for i, row_vals in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=start_data_row + i, column=col_idx, value=val)

    orders_count = estimate_unique_order_count(df)

    summary_row_idx = start_data_row + n_rows
    ws.cell(row=summary_row_idx, column=1, value="Orders on day:")
    ws.cell(row=summary_row_idx, column=2, value=orders_count)
    wb.save(xlsx_path)

# -----------------------------
# MONTHLY WORKBOOK (append-only, skip existing weeks)
# -----------------------------
def month_sheet_name(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"

def append_week_block(ws, df: pd.DataFrame, label: str, spacer_rows: int = 3):
    """
    Append a weekly block at the bottom of the sheet:
      - 1 row: "Week: YYYY.MM.DD-YYYY.MM.DD"
      - data rows (NO header)
      - 1 row: "Orders in week:" <unique innerId count>
      - spacer_rows empty rows
    """
    insert_at = (ws.max_row or 1) + 1

    # Label
    ws.cell(row=insert_at, column=1, value=f"Week: {label}")

    # Data rows
    for i, row_vals in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=insert_at + 1 + i, column=c_idx, value=val)

    # Summary
    orders_count = estimate_unique_order_count(df)
    summary_row = insert_at + 1 + len(df)
    ws.cell(row=summary_row, column=1, value="Orders in week:")
    ws.cell(row=summary_row, column=2, value=orders_count)

    # Spacer rows
    for idx in range(max(0, spacer_rows)):
        ws.cell(row=summary_row + idx + 1, column=2, value="")

def week_months_covered(start_dt: date, end_dt: date) -> list[date]:
    months = set()
    cur = start_dt

    while cur <= end_dt:
        months.add(date(cur.year, cur.month, 1))
        cur += timedelta(days=1)

    return sorted(months)

def weekly_ranges_between(start_dt: date, end_dt: date) -> list[tuple[date, date]]:
    first_monday = start_dt - timedelta(days=start_dt.weekday())
    ranges = []
    cur_start = first_monday

    while cur_start <= end_dt:
        cur_end = cur_start + timedelta(days=6)
        s = max(cur_start, start_dt)
        e = min(cur_end, end_dt)
        ranges.append((s, e))
        cur_start += timedelta(days=7)

    return ranges

def build_monthly_workbook_for_previous_weeks(access_token,
                                              token_type,
                                              months_back: int = 3,
                                              out_xlsx: str = os.path.join(DATA_DIR, "orders_popfanatic_by_month.xlsx"),
                                              spacer_rows: int = 3) -> str:
    """
    RUN WEEKLY (e.g., every Monday):
    - Append only the **missing** weeks to month sheets.
    - Never delete or rewrite existing weeks.
    - If a week touches two months, write to both sheets (but only if missing there).
    """
    ensure_data_dir()

    # Determine window [window_start .. today]
    today = date.today()
    cur_month_first = date(today.year, today.month, 1)
    start_year = cur_month_first.year
    start_month = cur_month_first.month - months_back

    while start_month <= 0:
        start_month += 12
        start_year -= 1

    window_start = date(start_year, start_month, 1)
    window_end = today

    # All week ranges in window
    weeks = weekly_ranges_between(window_start, window_end)

    # We will open the workbook lazily per sheet to read existing labels and header
    existing_wb = load_workbook(out_xlsx) if os.path.exists(out_xlsx) else None
    existing_labels_by_sheet: dict[str, set[str]] = {}
    existing_headers_by_sheet: dict[str, list[str]] = {}

    if existing_wb:
        for sheet in existing_wb.sheetnames:
            ws = existing_wb[sheet]
            existing_labels_by_sheet[sheet] = _existing_week_labels_in_sheet(ws)
            existing_headers_by_sheet[sheet] = _get_existing_header(ws)

    def sheet_has_label(sheet_name: str, label: str) -> bool:
        return label in existing_labels_by_sheet.get(sheet_name, set())

    # Iterate weeks; only fetch those that are missing from at least one affected month sheet
    for (ws_start, ws_end) in weeks:
        label = f"{ws_start.strftime('%Y.%m.%d')}-{ws_end.strftime('%Y.%m.%d')}"
        months_hit = [month_sheet_name(m) for m in week_months_covered(ws_start, ws_end)]

        # Skip if week already exists on ALL relevant sheets
        if months_hit and all(sheet_has_label(sheet, label) for sheet in months_hit):
            continue

        # Fetch details for this week (only now, since it's missing somewhere)
        start_iso = datetime(ws_start.year, ws_start.month, ws_start.day, 0, 0, 0).strftime("%Y-%m-%dT%H:%M:%S")
        end_iso   = datetime(ws_end.year, ws_end.month, ws_end.day, 23, 59, 59).strftime("%Y-%m-%dT%H:%M:%S")
        stubs = fetch_orders_between(access_token, token_type, start_iso, end_iso)

        if not stubs:
            # Even if missing, nothing to write
            continue

        details = fetch_order_details_for_items(access_token, token_type, stubs)
        df_week = details_to_dataframe(details)

        # Write this week to every month sheet where it's missing
        for sheet in months_hit:
            # Header for this sheet: preserve existing header if any, else derive from df
            header_cols = existing_headers_by_sheet.get(sheet)
            if not header_cols or all(h is None for h in header_cols):
                header_cols = sorted(df_week.columns) if not df_week.empty else ["info"]

            # Align df to header (drop unknown cols; add missing empty cols)
            df_aligned = df_week.reindex(columns=header_cols)

            wb, ws = _open_or_init_wb_with_header(out_xlsx, sheet, header_cols)

            # Refresh caches if we just created a new sheet in a new file
            if sheet not in existing_labels_by_sheet:
                existing_labels_by_sheet[sheet] = _existing_week_labels_in_sheet(ws)
                existing_headers_by_sheet[sheet] = _get_existing_header(ws)

            # Only append if missing on this particular sheet
            if label not in existing_labels_by_sheet[sheet]:
                append_week_block(ws, df_aligned, label=label, spacer_rows=spacer_rows)
                wb.save(out_xlsx)
                # Update cache so repeated weeks in same run won’t duplicate
                existing_labels_by_sheet[sheet].add(label)

    print(f"Monthly workbook ready (append-only): {out_xlsx}")
    return out_xlsx

# -----------------------------
# DAILY SUMMARY (prepend TODAY, then YESTERDAY)
# -----------------------------
def daily_summary_orders_into_excel(access_token, token_type,
                                   output_path: str = DEFAULT_MAIN_XLSX,
                                   sheet_name: str = DEFAULT_SHEET,
                                   spacer_rows: int = 3) -> None:
    """
    Rolling daily summary:
    - remove any previous partial Yesterday,
    - prepend Yesterday (final) then Today (current),
    - summary counts by unique innerId.
    """
    ensure_data_dir()

    now = datetime.now()
    today_date_str = now.strftime("%Y-%m-%d")
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).strftime("%Y-%m-%dT%H:%M:%S")
    today_end   = now.strftime("%Y-%m-%dT%H:%M:%S")

    y = now - timedelta(days=1)
    yday_date_str = y.strftime("%Y-%m-%d")
    yday_start = y.replace(hour=0, minute=0, second=0, microsecond=0).strftime("%Y-%m-%dT%H:%M:%S")
    yday_end   = y.replace(hour=23, minute=59, second=59, microsecond=0).strftime("%Y-%m-%dT%H:%M:%S")

    today_stubs = fetch_orders_between(access_token, token_type, today_start, today_end)
    yday_stubs  = fetch_orders_between(access_token, token_type, yday_start, yday_end)

    today_details = fetch_order_details_for_items(access_token, token_type, today_stubs) if today_stubs else []
    yday_details  = fetch_order_details_for_items(access_token, token_type, yday_stubs)  if yday_stubs  else []

    df_today = details_to_dataframe(today_details)
    df_yday  = details_to_dataframe(yday_details)

    header_cols = sorted(set(df_today.columns).union(df_yday.columns)) or ["info"]
    df_today = df_today.reindex(columns=header_cols)
    df_yday  = df_yday.reindex(columns=header_cols)

    _open_or_init_wb_with_header(output_path, sheet_name, header_cols)[0].save(output_path)

    deleted = delete_batch_by_label(output_path, sheet_name, yday_date_str, header_cols)
    if deleted:
        print(f"• Removed previous partial day for {yday_date_str}")

    prepend_batch_to_excel(df_yday,  output_path, batch_label=yday_date_str,  sheet_name=sheet_name, spacer_rows=spacer_rows)
    prepend_batch_to_excel(df_today, output_path, batch_label=today_date_str, sheet_name=sheet_name, spacer_rows=spacer_rows)

    print(f"✔ Rotated batches. Top = TODAY({today_date_str}), below = YESTERDAY({yday_date_str}). File: {output_path}")

# -----------------------------
# Optional: write all orders of today into per-day files
# -----------------------------
def get_today_orders_write_into_excel(access_token, token_type) -> None:
    ensure_data_dir()

    today = datetime.now().strftime("%Y-%m-%d")
    extra_params = {"createdAt": today}

    data = get_orders(access_token, token_type, page=0, limit=200, extra_params=extra_params)
    ndjson_path = os.path.join(DATA_DIR, f"today_popfanatic_{today}.ndjson")
    xlsx_path = os.path.join(DATA_DIR, f"today_popfanatic_{today}.xlsx")

    items = data.get("items") or []
    if len(items) > 0:
        with open(ndjson_path, "w", encoding="utf-8") as f_out:
            for item in items:
                id_ = extract_order_id(item)
                order = get_order_by_id(access_token, token_type, id_)
                f_out.write(json.dumps(order, ensure_ascii=False) + "\n")

        df = pd.read_json(ndjson_path, lines=True)
        df.to_excel(xlsx_path, index=False, engine="openpyxl")
    else:
        pd.DataFrame([{"orders": 0, "createdAt": today}]).to_excel(xlsx_path, index=False, engine="openpyxl")

    print("Today orders written to excel")


# -----------------------------
# Main
# -----------------------------
def main() -> None:
    access_token, token_type = get_access_token()

    # Run weekly (e.g., every Monday): append-only, skip existing weeks
    build_monthly_workbook_for_previous_weeks(
        access_token=access_token,
        token_type=token_type,
        months_back=3,
        out_xlsx=os.path.join(DATA_DIR, "orders_popfanatic_by_month.xlsx"),
        spacer_rows=10
    )

    # Daily rolling summary with top-insert (TODAY then YESTERDAY)
    daily_summary_orders_into_excel(
        access_token=access_token,
        token_type=token_type,
        output_path=DEFAULT_MAIN_XLSX,
        sheet_name=DEFAULT_SHEET,
        spacer_rows=3
    )

    # Get today orders
    get_today_orders_write_into_excel(
        access_token=access_token,
        token_type=token_type
    )

if __name__ == "__main__":
    main()
