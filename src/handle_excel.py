import os
import logging
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from dotenv import load_dotenv
from openpyxl.workbook import Workbook
import pandas as pd
import tempfile

# =====================================================
# Logging setup
# =====================================================
# Create log folder if missing
LOG_DIR = os.path.join(os.path.dirname(__file__), "logs")
os.makedirs(LOG_DIR, exist_ok=True)

# Create log file (append mode, one file per day)
log_file = os.path.join(LOG_DIR, f"excel_processing_{datetime.now().strftime('%Y-%m-%d')}.log")

logging.basicConfig(
    level=logging.INFO,  # INFO = normal messages; use DEBUG for more verbosity
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[
        logging.FileHandler(log_file, mode="a", encoding="utf-8"),  # append instead of overwrite
        logging.StreamHandler()  # show logs also in terminal
    ]
)

logger = logging.getLogger(__name__)
logger.info("Logging initialized -> %s", log_file)

# =====================================================
# Load environment
# =====================================================
load_dotenv()
download_folder = os.getenv("DOWNLOAD_DIR")

# =====================================================
# Helpers
# =====================================================
def rename_excel_sheet(wb: Workbook, new_sheet_name: str, path: str) -> None:
    sheet = wb[wb.sheetnames[0]]
    sheet.title = new_sheet_name
    wb.save(path)


def rename_excel(excel_path: str, to: str) -> None:
    os.rename(excel_path, to)


def filter_excel(excel_path: str, webshop_name: str) -> None:
    """
    Keep only required columns and ensure 'Termék egységára' exists.
    - Renames various 'nettó ár' variants to 'Termék egységára'.
    - If no direct unit-price column exists, computes it from 'Nettó Összesen' / 'Mennyiség' when available.
    """
    import unicodedata
    import pandas as pd
    import numpy as np

    def norm(s: str) -> str:
        if s is None:
            return ""
        s = unicodedata.normalize("NFKC", str(s))
        s = s.replace("\u00A0", " ")  # non-breaking space -> normal space
        return s.strip().lower()

    desired_cols = [
        "Rendelés szám", "Vásárló csoport", "E-mail", "Dátum",
        "Szállítási Mód", "Fizetési Mód", "Megrendelés státusz",
        "Száll. Város", "Száll. Ir.", "Száll. Ország", "Adószám",
        "Szállítási Díj", "Kezelési Költség", "Kedvezmény",
        "Termék Név", "Cikkszám", "Mennyiség", "Termék egységára",
        "Nettó Összesen"
    ]

    unit_price_aliases = {
        "nettó ár", "netto ár", "netto ar", "nettó ar",
        "nettó egységár", "egységár (nettó)", "egysegar (netto)",
        "unit net price", "net unit price", "unit price (net)"
    }

    df = pd.read_excel(excel_path)
    norm_to_orig = {norm(c): c for c in df.columns}

    # 1️⃣ Find a column that matches known unit price names
    unit_col_orig = None
    for alias in unit_price_aliases:
        if alias in norm_to_orig:
            unit_col_orig = norm_to_orig[alias]
            break

    # 2️⃣ Rename or compute 'Termék egységára'
    if unit_col_orig:
        if unit_col_orig != "Termék egységára":
            df.rename(columns={unit_col_orig: "Termék egységára"}, inplace=True)
    else:
        sum_col = norm_to_orig.get("nettó összesen") or norm_to_orig.get("netto osszesen") or norm_to_orig.get("netto összesen")
        qty_col = norm_to_orig.get("mennyiség") or norm_to_orig.get("mennyiseg")
        if sum_col and qty_col:
            with np.errstate(all="ignore"):
                df["Termék egységára"] = pd.to_numeric(df[sum_col], errors="coerce") / pd.to_numeric(df[qty_col], errors="coerce")
        else:
            df["Termék egységára"] = np.nan

    # 3️⃣ Keep only desired columns that exist
    keep = [c for c in desired_cols if c in df.columns]
    df = df[keep]

    # 4️⃣ Replace empty 'Vásárló csoport' with '|'
    vc = "Vásárló csoport"
    if vc in df.columns:
        mask = df[vc].isna() | (df[vc].astype(str).str.strip() == "")
        df.loc[mask, vc] = "|"

    if "Dátum" in df.columns:
        df["Dátum"] = pd.to_datetime(df["Dátum"], errors="coerce").dt.strftime("%Y.%m").str.replace(r"\.([0-9])$",
                                                                                                    r".0\1", regex=True)

    df["Bolt neve"] = webshop_name
    df.to_excel(excel_path, index=False, float_format="%.2f")
    logger.info("Filtered and saved: %s", excel_path)


def summarize_orders_into_excel(path: str) -> pd.DataFrame:
    """Summarize each day’s Excel file (order count + total revenue)."""
    df = pd.read_excel(path)
    rows = len(df)

    wb = load_workbook(path)
    day_name = wb.active.title  # usually 'YYYY-MM-DD'

    cols_net = ["Nettó Összesen", "Kedvezmény"]
    cols_gross = ["Szállítási Díj", "Kezelési Költség"]
    vat_multiplier = 0.73

    net_sum = df[cols_net].sum().sum() if all(col in df.columns for col in cols_net) else 0
    gross_sum = (df[cols_gross].sum().sum() * vat_multiplier) if all(col in df.columns for col in cols_gross) else 0
    total_revenue = net_sum + gross_sum

    out = pd.DataFrame({
        day_name: [rows, total_revenue],
        "": ["", ""]
    }, index=["Orders", "Revenue"])

    # Force text with '.' instead of ','
    out = out.applymap(lambda x: f"{x:.2f}" if isinstance(x, (int, float)) else x)

    return out


def merge_all_daily_summaries(folder_path: str) -> None:
    """Combine all 'day-*.xlsx' summaries into a single daily-summary.xlsx."""
    files = sorted(
        [f for f in os.listdir(folder_path) if f.lower().endswith(".xlsx") and f.startswith("day-")]
    )
    if not files:
        return

    all_data = pd.DataFrame()
    for file in files:
        file_path = os.path.join(folder_path, file)
        df_summary = summarize_orders_into_excel(file_path)
        all_data = pd.concat([all_data, df_summary], axis=1)

    out_path = os.path.join(folder_path, "daily-summary.xlsx")
    all_data.to_excel(out_path, float_format="%.2f")
    logger.info("✅ Saved merged summary to %s", out_path)


# =====================================================
# NEW: temp builder + incremental sync
# =====================================================
def _atomic_write_xlsx(df: pd.DataFrame, out_path: str) -> None:
    tmp_fd, tmp_path = tempfile.mkstemp(prefix="daily-summary-", suffix=".xlsx",
                                        dir=os.path.dirname(out_path) or ".")
    os.close(tmp_fd)
    try:
        df.to_excel(tmp_path, float_format="%.2f")
        os.replace(tmp_path, out_path)
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


def build_daily_summary_temp_df(folder_path: str) -> pd.DataFrame:
    """
    Build a temporary DataFrame from all day-*.xlsx in folder_path.
    Also writes it to 'daily_summary_temp.xlsx' for inspection.
    """
    day_files = sorted(
        [f for f in os.listdir(folder_path)
         if f.lower().endswith(".xlsx") and f.startswith("day-")]
    )
    all_data = pd.DataFrame()
    for f in day_files:
        fp = os.path.join(folder_path, f)
        col = summarize_orders_into_excel(fp)
        all_data = pd.concat([all_data, col], axis=1)

    # Save temp file
    temp_path = os.path.join(folder_path, "daily_summary_temp.xlsx")
    _atomic_write_xlsx(all_data, temp_path)
    logger.info("🧪 Wrote temporary daily summary to %s", temp_path)

    return all_data


def sync_daily_summary_from_temp(folder_path: str) -> None:
    """
    If daily-summary.xlsx exists, create daily_summary_temp.xlsx and append any missing
    day columns (by date) into daily-summary.xlsx (with a blank spacer after each new day).
    If daily-summary.xlsx doesn't exist, create it from all day-*.xlsx files.
    """
    out_path = os.path.join(folder_path, "daily-summary.xlsx")

    temp_df = build_daily_summary_temp_df(folder_path)

    if not os.path.exists(out_path):
        # No summary yet → write full temp as the initial summary
        _atomic_write_xlsx(temp_df, out_path)
        logger.info("🆕 Created daily-summary.xlsx from temp in %s", folder_path)
        return

    # Load existing summary (keep index so 'Orders/Revenue' align)
    existing = pd.read_excel(out_path, index_col=0)

    # Identify day columns in temp (exclude blank spacer column name "")
    temp_cols = list(temp_df.columns)
    day_cols = [c for c in temp_cols if c != ""]

    # Determine which day columns are missing from existing
    missing_days = [d for d in day_cols if d not in existing.columns]

    if not missing_days:
        logger.info("ℹ️ No new days to append in %s", folder_path)
        return

    # Build a DataFrame to append: for each missing day, add the day column and a new blank spacer
    pieces = []
    for d in missing_days:
        day_part = temp_df[[d]]
        spacer = pd.DataFrame({"": ["", ""]}, index=day_part.index)
        pair = pd.concat([day_part, spacer], axis=1)
        pieces.append(pair)

    to_append = pd.concat(pieces, axis=1) if pieces else pd.DataFrame(index=existing.index)

    combined = pd.concat([existing, to_append], axis=1)
    _atomic_write_xlsx(combined, out_path)
    logger.info("➕ Appended %d new day(s) to %s", len(missing_days), out_path)


# =====================================================
# Main workflow
# =====================================================
def move_files_into_webshop_folders() -> None:
    """Move all downloaded .xlsx files into subfolders by webshop name."""
    for file in os.listdir(download_folder):
        if not file.lower().endswith(".xlsx"):
            continue

        try:
            webshop_name_local = file.split("_")[1].split("-")[0]
            if webshop_name_local == "tesztpr":
                webshop_name_local = "jatekfarm"
            if webshop_name_local == "toymarket":
                webshop_name_local = "tarsasjatekrendeles"
        except Exception:
            webshop_name_local = "unknown"

        folder_path = os.path.join(download_folder, webshop_name_local)
        os.makedirs(folder_path, exist_ok=True)

        src = os.path.join(download_folder, file)
        dst = os.path.join(folder_path, file)
        if os.path.abspath(src) != os.path.abspath(dst):
            os.replace(src, dst)

    today = date.today()

    # Rename daily files within each webshop folder
    for folder in os.listdir(download_folder):
        folder_path = os.path.join(download_folder, folder)
        if not os.path.isdir(folder_path):
            continue

        xlsx_files = [
            f for f in os.listdir(folder_path)
            if f.lower().endswith(".xlsx")
            and not f.startswith("day-")
            and not f.startswith("year-")
            and f.lower() != "daily-summary.xlsx"
        ]
        if not xlsx_files:
            # still try to sync summary (in case day- files already present)
            sync_daily_summary_from_temp(folder_path)
            continue

        xlsx_files.sort(key=lambda f: os.path.getmtime(os.path.join(folder_path, f)), reverse=True)

        for idx, fname in enumerate(xlsx_files):
            target_date = today - timedelta(days=idx - 1)
            base_date = target_date.isoformat()
            new_base = f"day-{base_date}"
            new_name = f"{new_base}.xlsx"

            src = os.path.join(folder_path, fname)
            dst = os.path.join(folder_path, new_name)

            # Ensure unique names if duplicates exist
            final_dst = dst
            if os.path.exists(final_dst):
                i = 1
                while True:
                    candidate = os.path.join(folder_path, f"{new_base}_{i}.xlsx")
                    if not os.path.exists(candidate):
                        final_dst = candidate
                        break
                    i += 1

            os.rename(src, final_dst)

            try:
                wb = load_workbook(final_dst)
                rename_excel_sheet(wb, base_date, final_dst)
            except Exception as e:
                logger.warning("⚠️ Sheet rename skipped for %s: %s", final_dst, e)

            logger.info("📦 %s: '%s' → '%s'", folder, fname, os.path.basename(final_dst))

        # Create yearly summary (largest file)
        day_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".xlsx") and f.startswith("day-")]
        if day_files:
            largest_file = max(day_files, key=lambda f: os.path.getsize(os.path.join(folder_path, f)))
            largest_path = os.path.join(folder_path, largest_file)
            year_name = f"year-{today.year}.xlsx"
            year_path = os.path.join(folder_path, year_name)

            if os.path.abspath(largest_path) != os.path.abspath(year_path):
                if os.path.exists(year_path):
                    os.remove(year_path)
                os.rename(largest_path, year_path)

                filter_excel(year_path, folder)

                try:
                    wb = load_workbook(year_path)
                    rename_excel_sheet(wb, str(today.year), year_path)
                except Exception as e:
                    logger.warning("⚠️ Year sheet rename skipped for %s: %s", year_path, e)
                logger.info("🏷️ %s: '%s' → '%s'", folder, os.path.basename(largest_path), year_name)

        # 🔁 NEW: Incremental sync using temp vs existing
        sync_daily_summary_from_temp(folder_path)


def delete_unnecessary_files(download_dir: str) -> None:
    """Remove temporary 'day-' Excel files after merging."""
    for folder in os.listdir(download_dir):
        folder_path = os.path.join(download_dir, folder)
        if not os.path.isdir(folder_path):
            continue
        for file in os.listdir(folder_path):
            if file.startswith("day-") and file.lower().endswith(".xlsx"):
                os.remove(os.path.join(folder_path, file))
                logger.info("🗑️ Deleted temporary file: %s/%s", folder, file)

            if file.find('temp') != -1 and file.endswith('.xlsx'):
                os.remove(os.path.join(folder_path, file))
                logger.info("Deleted temporary file: %s/%s", folder, file)


def main() -> None:
    logger.info("=== Starting Excel file organization ===")
    move_files_into_webshop_folders()
    delete_unnecessary_files(download_folder)
    logger.info("✅ Excel organization complete.")


if __name__ == "__main__":
    main()
